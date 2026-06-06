from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models import ProcessEmailRequest, ApproveRequest, AmendRequest, Deal
from app.extraction import extract_deal_data
from app.emailer import send_review_notification
from app.database import get_all_deals, get_deal_by_id, save_deal, update_deal, delete_deal
from dotenv import load_dotenv

load_dotenv()

# Ensure directories exist on startup
os.makedirs("app/data", exist_ok=True)
os.makedirs("app/outputs", exist_ok=True)

app = FastAPI(
    title="Electric Land Valuation Agent API",
    description="Automates deal data extraction and valuation model population",
    version="1.0.0"
)

# Allow frontend to talk to this API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

OUTPUTS_DIR = "app/outputs"# Create directories if they don't exist
os.makedirs("app/data", exist_ok=True)
os.makedirs("app/outputs", exist_ok=True)#hi
# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Generate Excel valuation model
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(deal: dict) -> str:
    d = deal["extracted_data"]

    dark_green  = "1A3C2E"
    mid_green   = "2D6A4F"
    light_green = "D8F3DC"
    amber       = "F59E0B"
    white       = "FFFFFF"
    light_grey  = "F8FAFC"
    mid_grey    = "E2E8F0"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Model"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22

    def style(row, col, value, bold=False, bg=None, fc="000000",
              size=11, align="left", border=False, italic=False, fmt=None, merge_to=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fc, size=size, name="Calibri", italic=italic)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if border:
            s = Side(style="thin", color="CBD5E1")
            c.border = Border(left=s, right=s, top=s, bottom=s)
        if fmt:
            c.number_format = fmt
        return c

    def section(row, title):
        ws.row_dimensions[row].height = 26
        ws.merge_cells(f"A{row}:E{row}")
        style(row, 1, f"  {title}", bold=True, bg=mid_green, fc=white, size=12)

    def row_data(row, label, val, conf=None, is_currency=False, is_percent=False):
        ws.row_dimensions[row].height = 20
        bg = light_grey if row % 2 == 0 else white
        style(row, 1, label, bg=bg, border=True)

        if val is None:
            c = style(row, 2, "— Not provided", bg=bg, border=True, fc="94A3B8")
        elif is_currency and isinstance(val, (int, float)):
            c = style(row, 2, val, bg=bg, border=True, fmt='£#,##0')
        elif is_percent and isinstance(val, (int, float)):
            c = style(row, 2, val / 100, bg=bg, border=True, fmt='0.00%')
        else:
            c = style(row, 2, val, bg=bg, border=True)

        if conf:
            conf_bg = {"high": "DCFCE7", "medium": "FEF3C7", "low": "FEE2E2"}
            conf_fc = {"high": "16A34A", "medium": "D97706", "low": "DC2626"}
            style(row, 3, f"● {str(conf).upper()}", bg=conf_bg.get(conf, white),
                  fc=conf_fc.get(conf, "000000"), size=9, bold=True, align="center", border=True)
        else:
            style(row, 3, "", bg=bg, border=True)

        style(row, 4, "", bg=bg, border=True)
        style(row, 5, "", bg=bg, border=True)

    # Header
    ws.row_dimensions[1].height = 45
    ws.merge_cells("A1:E1")
    style(1, 1, "ELECTRIC LAND — DEAL VALUATION MODEL",
          bold=True, bg=dark_green, fc=white, size=16, align="center")

    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:C2")
    style(2, 1, "⚠  DRAFT — PENDING HUMAN REVIEW & APPROVAL",
          bold=True, bg=amber, fc=white, size=11, align="center")
    ws.merge_cells("D2:E2")
    style(2, 4, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
          bg=mid_grey, size=10, align="right", italic=True)

    ws.row_dimensions[3].height = 8

    # Column headers
    ws.row_dimensions[4].height = 22
    for col, label in enumerate(["Field", "Extracted Value", "Confidence", "Analyst Override", "Notes"], 1):
        style(4, col, label, bold=True, bg=mid_grey, fc="475569", size=10, align="center", border=True)

    # Site info
    section(5, "📍  SITE INFORMATION")
    row_data(6,  "Site Address",        d.get("site_address", {}).get("value"),        d.get("site_address", {}).get("confidence"))
    row_data(7,  "Site Area (acres)",   d.get("site_area_acres", {}).get("value"),     d.get("site_area_acres", {}).get("confidence"))
    row_data(8,  "Planning Status",     d.get("planning_status", {}).get("value"),     d.get("planning_status", {}).get("confidence"))
    row_data(9,  "Dev. Timeline",
             f"{d.get('development_timeline_months', {}).get('value')} months"
             if d.get('development_timeline_months', {}).get('value') else None,
             d.get("development_timeline_months", {}).get("confidence"))
    row_data(10, "Deal Sender",         d.get("sender_name", {}).get("value"),         d.get("sender_name", {}).get("confidence"))

    # Financials
    section(11, "💷  FINANCIAL SUMMARY")
    row_data(12, "Purchase Price",         d.get("purchase_price_gbp", {}).get("value"),         d.get("purchase_price_gbp", {}).get("confidence"),         is_currency=True)
    row_data(13, "Build Cost",             d.get("build_cost_gbp", {}).get("value"),             d.get("build_cost_gbp", {}).get("confidence"),             is_currency=True)
    row_data(14, "Total Development Cost", d.get("total_development_cost_gbp", {}).get("value"), d.get("total_development_cost_gbp", {}).get("confidence"), is_currency=True)
    row_data(15, "GDV",                    d.get("gdv_gbp", {}).get("value"),                    d.get("gdv_gbp", {}).get("confidence"),                    is_currency=True)

    # Metrics
    section(16, "📊  CALCULATED METRICS")
    gdv = d.get("gdv_gbp", {}).get("value")
    tdc = d.get("total_development_cost_gbp", {}).get("value")
    profit_val = (gdv - tdc) if gdv and tdc else None
    row_data(17, "Profit on Cost (£)",  profit_val, is_currency=True)
    row_data(18, "Profit on Cost (%)",  d.get("profit_on_cost_percent", {}).get("value"),  d.get("profit_on_cost_percent", {}).get("confidence"),  is_percent=True)
    row_data(19, "Target Net Yield (%)",d.get("target_yield_percent", {}).get("value"),    d.get("target_yield_percent", {}).get("confidence"),    is_percent=True)

    # Rent roll
    section(20, "🏗️  RENT ROLL")
    row_data(21, "Total Rent Roll (p.a.)", d.get("total_rent_roll_pa_gbp", {}).get("value"), d.get("total_rent_roll_pa_gbp", {}).get("confidence"), is_currency=True)
    r = 22
    units = d.get("individual_units", {}).get("value") or []
    for unit in units:
        row_data(r, f"  └ {unit.get('unit_name','')}", unit.get("annual_rent_gbp"),
                 d.get("individual_units", {}).get("confidence"), is_currency=True)
        r += 1

    # Flags
    flags = d.get("flags", [])
    section(r, "⚑  AI FLAGS & WARNINGS")
    r += 1
    if flags:
        for flag in flags:
            ws.row_dimensions[r].height = 30
            ws.merge_cells(f"A{r}:E{r}")
            style(r, 1, f"⚠  {flag}", bg="FEF3C7", fc="92400E", size=10, border=True)
            r += 1
    else:
        ws.merge_cells(f"A{r}:E{r}")
        style(r, 1, "✅  No flags raised", bg=light_green, fc=mid_green, size=10, border=True)
        r += 1

    # Approval section
    r += 1
    section(r, "✅  HUMAN REVIEW & APPROVAL")
    r += 1
    for label in ["Reviewed by", "Review Date", "Approval Status", "Sign-off Notes"]:
        ws.row_dimensions[r].height = 22
        style(r, 1, label, bg=light_grey, border=True)
        ws.merge_cells(f"B{r}:E{r}")
        style(r, 2, "", border=True)
        r += 1

    # Footer
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    style(r, 1,
          "CONFIDENTIAL — Auto-generated draft. Must be reviewed and approved before use.",
          bg=dark_green, fc=white, size=9, align="center", italic=True)

    # Save
    filename = f"valuation_{deal['id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUTS_DIR, filename)
    wb.save(filepath)
    return filename


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {
        "name": "Electric Land Valuation Agent API",
        "version": "1.0.0",
        "status": "running",
        "endpoints": [
            "POST /process-email",
            "GET  /deals",
            "GET  /deals/{id}",
            "PATCH /deals/{id}/approve",
            "PATCH /deals/{id}/amend",
            "GET  /deals/{id}/download",
            "DELETE /deals/{id}"
        ]
    }


@app.post("/process-email")
def process_email(request: ProcessEmailRequest):
    """
    Takes an email, extracts deal data with Claude,
    generates Excel, saves deal, sends analyst notification.
    """
    # Extract with Claude
    try:
        extracted = extract_deal_data(request.body)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

    # Create deal record
    deal = Deal(
        email_subject=request.subject,
        email_from=request.email_from,
        email_body=request.body,
        extracted_data=extracted,
        status="pending_review"
    )

    # Generate Excel
    try:
        deal_dict = deal.model_dump()
        filename = generate_excel(deal_dict)
        deal.excel_filename = filename
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")

    # Save to database
    saved = save_deal(deal)

    # Send email notification
    try:
        send_review_notification(saved.model_dump(), os.path.join(OUTPUTS_DIR, filename))
    except Exception as e:
        print(f"Warning: Email notification failed: {e}")

    return {
        "success": True,
        "deal_id": saved.id,
        "status": saved.status,
        "site_address": extracted.site_address.value,
        "flags_raised": len(extracted.flags),
        "excel_filename": filename,
        "message": f"Deal processed successfully. Notification sent to analyst."
    }


@app.get("/deals")
def list_deals():
    """Returns all deals with summary info."""
    deals = get_all_deals()
    summary = []
    for d in deals:
        ed = d.get("extracted_data", {})
        summary.append({
            "id": d["id"],
            "created_at": d["created_at"],
            "status": d["status"],
            "email_subject": d["email_subject"],
            "email_from": d["email_from"],
            "site_address": ed.get("site_address", {}).get("value", "—"),
            "purchase_price": ed.get("purchase_price_gbp", {}).get("value"),
            "gdv": ed.get("gdv_gbp", {}).get("value"),
            "flags_count": len(ed.get("flags", [])),
            "excel_filename": d.get("excel_filename", "")
        })
    return {"total": len(summary), "deals": summary}


@app.get("/deals/{deal_id}")
def get_deal(deal_id: str):
    """Returns full deal data including all extracted fields."""
    deal = get_deal_by_id(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    return deal


@app.patch("/deals/{deal_id}/approve")
def approve_deal(deal_id: str, request: ApproveRequest):
    """Analyst approves the deal valuation."""
    deal = get_deal_by_id(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    updated = update_deal(deal_id, {
        "status": "approved",
        "approved_by": request.approved_by,
        "approved_at": datetime.now().strftime("%d %b %Y %H:%M:%S"),
        "analyst_notes": request.notes
    })

    return {
        "success": True,
        "deal_id": deal_id,
        "status": "approved",
        "approved_by": request.approved_by,
        "message": "Deal approved and signed off successfully."
    }


@app.patch("/deals/{deal_id}/amend")
def amend_deal(deal_id: str, request: AmendRequest):
    """Analyst amends extracted values and marks as amended."""
    deal = get_deal_by_id(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    # Apply amendments to extracted data
    extracted = deal.get("extracted_data", {})
    for field, new_value in request.amendments.items():
        if field in extracted:
            extracted[field]["value"] = new_value
            extracted[field]["confidence"] = "amended"

    updated = update_deal(deal_id, {
        "status": "amended",
        "extracted_data": extracted,
        "approved_by": request.amended_by,
        "approved_at": datetime.now().strftime("%d %b %Y %H:%M:%S"),
        "analyst_notes": request.notes
    })

    return {
        "success": True,
        "deal_id": deal_id,
        "status": "amended",
        "fields_amended": list(request.amendments.keys()),
        "message": "Deal amended and saved successfully."
    }


@app.get("/deals/{deal_id}/download")
def download_excel(deal_id: str):
    """Downloads the Excel valuation file for a deal."""
    deal = get_deal_by_id(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    filename = deal.get("excel_filename", "")
    filepath = os.path.join(OUTPUTS_DIR, filename)

    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Excel file not found")

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.delete("/deals/{deal_id}")
def delete_deal_route(deal_id: str):
    """Deletes a deal record."""
    success = delete_deal(deal_id)
    if not success:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    return {"success": True, "message": f"Deal {deal_id} deleted."}